import os
import re
import uuid
import httpx
from datetime import datetime
from fastapi import APIRouter, UploadFile, File, Form, Request, Response, HTTPException, Depends, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import pandas as pd
import io
import json
from app.services.ai_service import detect_columns, rewrite_message
from app.services.sheet_service import (
    read_sheet_data, 
    get_next_campaign_id,
    ensure_logs_sheet_headers,
    append_campaign_log,
    update_log_status,
    update_unsubscribe_status,
    read_users_data,
    check_unsubscribed_emails
)

router = APIRouter()

# In-memory storage for active sessions (token -> username)
ACTIVE_SESSIONS = {}

class LoginRequest(BaseModel):
    username: str
    password: str

async def get_current_user(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail="Unauthorized: Missing or invalid Authorization header token."
        )
    token = authorization.split(" ")[1]
    if token not in ACTIVE_SESSIONS:
        raise HTTPException(
            status_code=401,
            detail="Unauthorized: Session expired or invalid token."
        )
    return ACTIVE_SESSIONS[token]

@router.post("/login")
async def login(req: LoginRequest):
    # Authenticate using the external SQL login API on api.cubicalos.com
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.cubicalos.com/api/auth/login",
                json={
                    "user_id": req.username.strip(),
                    "password": req.password.strip()
                },
                headers={"Content-Type": "application/json"},
                timeout=8.0
            )
            
        if response.status_code == 200:
            res_data = response.json()
            if res_data.get("status") is True:
                # Successfully logged in via external database!
                token = res_data.get("access_token") or str(uuid.uuid4())
                user_obj = res_data.get("user") or {}
                username_val = user_obj.get("name") or req.username.strip()
                ext_roles = res_data.get("roles") or []
                role_val = "User"
                if ext_roles:
                    first_role_name = ext_roles[0].get("name", "User")
                    if "admin" in first_role_name.lower():
                        role_val = "Admin"
                    else:
                        role_val = first_role_name
                
                ACTIVE_SESSIONS[token] = {"username": username_val, "role": role_val}
                return {
                    "token": token,
                    "username": username_val,
                    "role": role_val
                }
            else:
                # Try fallback first before failing
                user_session = await _attempt_sheet_login_fallback(req.username, req.password)
                if user_session:
                    return user_session
                error_msg = res_data.get("message") or "Invalid User Name or Password."
                raise HTTPException(
                    status_code=401,
                    detail=f"Authentication failed: {error_msg}"
                )
        else:
            # Try fallback first
            user_session = await _attempt_sheet_login_fallback(req.username, req.password)
            if user_session:
                return user_session
            raise HTTPException(
                status_code=502,
                detail="External authentication service returned an error."
            )
            
    except HTTPException:
        raise
    except Exception as e:
        print(f"DEBUG: External login connection failed: {e}. Trying local sheet fallback...")
        user_session = await _attempt_sheet_login_fallback(req.username, req.password)
        if user_session:
            return user_session
        raise HTTPException(
            status_code=503,
            detail="Authentication service is currently offline or unreachable."
        )

async def _attempt_sheet_login_fallback(username: str, password: str):
    try:
        from app.services.sheet_service import read_users_data
        users_data = await read_users_data()
        if users_data and len(users_data) > 1:
            headers = [h.strip().lower() for h in users_data[0]]
            try:
                user_idx = headers.index("user name")
                pass_idx = headers.index("password")
                role_idx = headers.index("role")
            except ValueError:
                user_idx, pass_idx, role_idx = 1, 2, 3
            
            for row in users_data[1:]:
                if len(row) > max(user_idx, pass_idx):
                    sheet_user = row[user_idx].strip()
                    sheet_pass = row[pass_idx].strip()
                    sheet_role = row[role_idx].strip() if len(row) > role_idx else "User"
                    
                    if sheet_user.lower() == username.strip().lower() and sheet_pass == password.strip():
                        token = str(uuid.uuid4())
                        role_val = "Admin" if "admin" in sheet_role.lower() else "User"
                        ACTIVE_SESSIONS[token] = {"username": sheet_user, "role": role_val}
                        print(f"DEBUG: Local sheet login fallback SUCCESS for user: {sheet_user}")
                        return {
                            "token": token,
                            "username": sheet_user,
                            "role": role_val
                        }
    except Exception as sheet_err:
        print(f"DEBUG: Local sheet login fallback error: {sheet_err}")
    return None

@router.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    """
    Uploads a file, uses AI to detect columns, and returns a preview.
    """
    contents = await file.read()
    if file.filename.endswith('.csv'):
        df = pd.read_csv(io.BytesIO(contents))
    else:
        df = pd.read_excel(io.BytesIO(contents))
    
    df = df.fillna("")
    columns = df.columns.tolist()
    sample_data = df.head(3).to_dict(orient='records')
    
    # Detect columns using AI
    detected = await detect_columns(columns, sample_data)
    
    return {
        "filename": file.filename,
        "columns": columns,
        "sample_data": sample_data,
        "detected_mapping": detected
    }

import uuid
import asyncio
from fastapi import BackgroundTasks

# In-memory storage for campaign results (Replacement for Redis)
campaign_manager = {}

import time
# Simple cache for Google Sheets log fetches to reduce API calls and rate-limits
SHEETS_LOGS_CACHE = {}
CACHE_EXPIRY_SECONDS = 4  # short expiry to keep updates real-time but fast on multiple requests/polling

from app.services.whatsapp_service import send_whatsapp_message
from app.services.email_service import send_email

async def run_campaign_task(campaign_id, platform, spreadsheet_id, rows, headers, mapping_dict, whatsapp_message, email_subject, email_body, base_url, generated_by):
    print(f"\n{'='*40}")
    print(f"[CAMPAIGN START] ID: {campaign_id}")
    print(f"Total Rows: {len(rows)} | Platform: {platform}")
    print(f"{'='*40}")

    # Fetch unsubscribed email addresses
    unsubscribed_emails = set()
    try:
        unsubscribed_emails = await check_unsubscribed_emails(spreadsheet_id)
        print(f"INFO: Loaded {len(unsubscribed_emails)} unsubscribed email(s) from sheet logs.")
    except Exception as ex:
        print(f"DEBUG: Failed to check unsubscribed emails: {ex}")

    def get_idx(key, default_name):
        idx = headers.index(mapping_dict[key]) if mapping_dict.get(key) in headers else -1
        if idx == -1:
            for i, h in enumerate(headers):
                if h.lower() == default_name.lower(): return i
        return idx

    phone_idx = get_idx('phone', 'Phone')
    email_idx = get_idx('email', 'Email')
    name_idx = get_idx('name', 'Name')

    # Detect Company Name column — use AI mapping first, then keyword fallback
    company_idx = get_idx('company', 'Company Name')
    if company_idx == -1:
        # Extended keyword fallback for common column names
        company_candidates = ["company", "organization", "org", "firm", "business", "employer"]
        for i, h in enumerate(headers):
            if str(h).strip().lower() in company_candidates:
                company_idx = i
                break

    # Detect S NO column index for preserving original row numbers
    sno_idx = -1
    sno_candidates = ["s no", "s.no", "sno", "serial number", "sr no", "sr.no", "serialno", "s_no"]
    for i, h in enumerate(headers):
        if str(h).strip().lower() in sno_candidates:
            sno_idx = i
            break

    # Determine the public frontend URL for the unsubscribe page
    frontend_origin = os.getenv("FRONTEND_URL", "").rstrip("/")
    if not frontend_origin:
        backend_url = os.getenv("BACKEND_URL", "").rstrip("/")
        if backend_url:
            # Derive the frontend origin (strip /api if present)
            frontend_origin = backend_url.replace("api.", "").replace("/api", "")
        else:
            frontend_origin = base_url.rstrip("/")

    for i, row in enumerate(rows):
        row_index = i + 1
        name = row[name_idx] if (name_idx != -1 and name_idx < len(row)) else "Customer"
        company_name = row[company_idx] if (company_idx != -1 and company_idx < len(row)) else ""
        original_sno = str(row[sno_idx]).strip() if (sno_idx != -1 and sno_idx < len(row)) else str(row_index)
        print(f"-> [{i+1}/{len(rows)}] Processing: {name}")

        # Formatted timestamp for in-memory results
        sent_ts = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")

        # WhatsApp
        if platform in ["whatsapp", "both"]:
            clean_phone = ""
            if phone_idx == -1 or phone_idx >= len(row):
                result = {"status": "Failed", "reason": "WhatsApp column not found"}
            else:
                phone = str(row[phone_idx]).strip()
                clean_phone = phone.replace(" ", "").replace("-", "").strip("+")
                if not clean_phone:
                    result = {"status": "Failed", "reason": "Phone number is empty"}
                elif len(clean_phone) < 10:
                    result = {"status": "Failed", "reason": "Invalid phone number length"}
                else:
                    if len(clean_phone) == 10: clean_phone = "91" + clean_phone
                    msg = (whatsapp_message or "").replace("{name}", name)
                    try:
                        result = await send_whatsapp_message(clean_phone, msg)
                    except Exception as e:
                        print(f"   WhatsApp Error: {e}")
                        result = {"status": "Failed", "reason": str(e)}

            reason_str = f"{result['status']}: {result.get('reason')}" if result.get('reason') else result['status']
            campaign_manager[campaign_id]["results"].append({
                "campaign_id": campaign_id,
                "name": name,
                "company": company_name,
                "sno": original_sno,
                "phone": clean_phone,
                "type": "WhatsApp",
                "status": result['status'],
                "reason": reason_str,
                "row_index": row_index,
                "sent_time": sent_ts,
                "generated_by": generated_by,
                "subscription": "Yes",
                "unsub_reason": "",
                "unsub_other": ""
            })

            if result['status'] == "Sent":
                print(f"   WhatsApp: {result['status']}")
            else:
                print(f"   WhatsApp Failed: {result.get('reason', 'Unknown reason')}")

            await append_campaign_log(
                spreadsheet_id=spreadsheet_id,
                campaign_id=campaign_id,
                platform="WhatsApp",
                name=name,
                phone=clean_phone,
                email="",
                status=result['status'],
                details=result.get('reason', ''),
                generated_by=generated_by,
                company_name=company_name
            )

        # Email
        if platform in ["email", "both"]:
            email = ""
            if email_idx == -1 or email_idx >= len(row):
                result = {"status": "Failed", "reason": "Email column not found"}
            else:
                email = str(row[email_idx]).strip()
                if not email:
                    result = {"status": "Failed", "reason": "Email address is empty"}
                elif not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
                    result = {"status": "Failed", "reason": "Invalid email address format"}
                elif email.lower() in unsubscribed_emails:
                    result = {"status": "Unsubscribed", "reason": "Recipient previously unsubscribed"}
                else:
                    body = (email_body or "").replace("{name}", name)
                    tracking_url = f"{base_url}api/campaign/track-open?campaign_id={campaign_id}&email={email}"
                    # Build unsubscribe URL pointing to the frontend unsubscribe page
                    import urllib.parse
                    unsub_base_url = os.getenv("UNSUBSCRIBE_PAGE_URL", "https://unsubscribe.cubemoons.com/index.html")
                    if not unsub_base_url.startswith("http://") and not unsub_base_url.startswith("https://"):
                        unsub_base_url = f"{frontend_origin}/unsubscribe/index.html"
                    
                    unsubscribe_url = (
                        f"{unsub_base_url}"
                        f"?email={urllib.parse.quote(email)}"
                        f"&campaign_id={urllib.parse.quote(campaign_id)}"
                    )
                    try:
                        result = await send_email(
                            email,
                            email_subject or "Update",
                            body,
                            tracking_url=tracking_url,
                            unsubscribe_url=unsubscribe_url
                        )
                    except Exception as e:
                        print(f"   Email Error: {e}")
                        result = {"status": "Failed", "reason": str(e)}

            reason_str = f"{result['status']}: {result.get('reason')}" if result.get('reason') else result['status']
            is_unsub = result['status'] == "Unsubscribed"
            campaign_manager[campaign_id]["results"].append({
                "campaign_id": campaign_id,
                "name": name,
                "company": company_name,
                "sno": original_sno,
                "email": email,
                "type": "Email",
                "status": result['status'],
                "reason": reason_str,
                "row_index": row_index,
                "sent_time": sent_ts,
                "generated_by": generated_by,
                "subscription": "No" if is_unsub else "Yes",
                "unsub_reason": "Recipient previously unsubscribed" if is_unsub else "",
                "unsub_other": ""
            })

            if result['status'] == "Sent":
                print(f"   Email: {result['status']}")
            else:
                print(f"   Email Status: {result['status']} ({result.get('reason', 'Unknown reason')})")

            await append_campaign_log(
                spreadsheet_id=spreadsheet_id,
                campaign_id=campaign_id,
                platform="Email",
                name=name,
                phone="",
                email=email,
                status=result['status'],
                details=result.get('reason', ''),
                generated_by=generated_by,
                company_name=company_name,
                subscription="No" if is_unsub else "Yes"
            )

        await asyncio.sleep(2)
    print(f"\n[CAMPAIGN FINISHED] ID: {campaign_id}\n")

@router.post("/start-campaign")
async def start_campaign(
    background_tasks: BackgroundTasks,
    request: Request,
    platform: str = Form(...),
    spreadsheet_id: str = Form(None),
    whatsapp_message: str = Form(None),
    email_subject: str = Form(None),
    email_body: str = Form(None),
    mapping: str = Form(...),
    start_sno: str = Form(None),
    end_sno: str = Form(None),
    file: UploadFile = File(None),
    current_user: dict = Depends(get_current_user)
):
    mapping_dict = json.loads(mapping)
    
    if file is not None:
        try:
            contents = await file.read()
            if file.filename.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(contents))
            else:
                df = pd.read_excel(io.BytesIO(contents))
            
            df = df.fillna("")
            headers = df.columns.tolist()
            rows = []
            for r in df.values.tolist():
                rows.append([str(x) if x is not None and str(x) != 'nan' else "" for x in r])
        except Exception as e:
            print(f"DEBUG: Error reading uploaded file: {e}")
            raise HTTPException(
                status_code=400,
                detail=f"Failed to process uploaded file: {str(e)}"
            )
    else:
        if not spreadsheet_id:
            raise HTTPException(
                status_code=400,
                detail="Please provide a Google Sheet ID or upload a local file."
            )
        try:
            data = await read_sheet_data(spreadsheet_id)
        except Exception as e:
            print(f"DEBUG: Error reading spreadsheet data: {e}")
            raise HTTPException(
                status_code=400,
                detail="Failed to read Google Sheet. Please share it with the service account (audit-ai@recruiterai-492605.iam.gserviceaccount.com) as at least a Viewer."
            )
            
        if not data:
            raise HTTPException(
                status_code=400,
                detail="No data found in Google Sheet."
            )
        
        headers = data[0]
        rows = data[1:]
    
    await ensure_logs_sheet_headers(spreadsheet_id)
    
    # Parse S NO inputs
    start_sno_int = None
    end_sno_int = None
    if start_sno and start_sno.strip():
        try:
            start_sno_int = int(float(start_sno.strip()))
        except ValueError:
            raise HTTPException(status_code=400, detail="Start S.No must be a valid integer.")
    if end_sno and end_sno.strip():
        try:
            end_sno_int = int(float(end_sno.strip()))
        except ValueError:
            raise HTTPException(status_code=400, detail="End S.No must be a valid integer.")
            
    # Filter by S NO range if provided
    filtered_rows = []
    if start_sno_int is not None or end_sno_int is not None:
        sno_idx = -1
        sno_candidates = ["s no", "s.no", "sno", "serial number", "sr no", "sr.no", "serialno", "s_no"]
        for idx, header in enumerate(headers):
            if str(header).strip().lower() in sno_candidates:
                sno_idx = idx
                break
                
        if sno_idx == -1:
            raise HTTPException(
                status_code=400,
                detail="S NO column not found in Google Sheet. Please ensure your sheet contains an 'S NO' or 'S.No' column."
            )
            
        for r in rows:
            if sno_idx >= len(r):
                continue
            sno_val_str = str(r[sno_idx]).strip()
            if not sno_val_str:
                continue
            try:
                sno_val = int(float(sno_val_str))
            except ValueError:
                continue
                
            if start_sno_int is not None and sno_val < start_sno_int:
                continue
            if end_sno_int is not None and sno_val > end_sno_int:
                continue
                
            filtered_rows.append(r)
            
        if not filtered_rows:
            raise HTTPException(
                status_code=400,
                detail=f"No rows found with S NO in the range {start_sno_int or ''} to {end_sno_int or ''}."
            )
    else:
        filtered_rows = rows
        
    campaign_base_id = await get_next_campaign_id(spreadsheet_id)
    if start_sno_int is not None and end_sno_int is not None:
        campaign_id = f"{campaign_base_id}({start_sno_int}-{end_sno_int})"
    elif start_sno_int is not None:
        campaign_id = f"{campaign_base_id}({start_sno_int}-)"
    elif end_sno_int is not None:
        campaign_id = f"{campaign_base_id}(-{end_sno_int})"
    else:
        campaign_id = campaign_base_id
        
    campaign_manager[campaign_id] = {
        "metadata": {
            "spreadsheet_id": spreadsheet_id,
            "platform": platform,
            "created_by": current_user["username"]
        },
        "results": []
    }
    
    base_url = os.getenv("BACKEND_URL") or str(request.base_url)
    if not base_url.endswith("/"):
        base_url += "/"
    print(f"DEBUG: Queueing campaign task for {len(filtered_rows)} contacts...")
    background_tasks.add_task(
        run_campaign_task, 
        campaign_id, platform, spreadsheet_id, filtered_rows, headers, mapping_dict, whatsapp_message, email_subject, email_body, base_url, current_user["username"]
    )
    
    return {
        "message": "Campaign started successfully", 
        "campaign_id": campaign_id, 
        "total_contacts": len(filtered_rows)
    }


import time

EMAILS_SENT_TODAY_CACHE = {
    "count": 0,
    "timestamp": 0.0
}

def fetch_gmail_sent_count(smtp_user, smtp_pass, smtp_server):
    import imaplib
    import re
    from datetime import datetime, timedelta
    
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(smtp_user, smtp_pass)
    
    sent_folder = None
    status, folder_list = mail.list()
    if status == 'OK':
        for folder_info in folder_list:
            folder_str = folder_info.decode('utf-8')
            if '\\sent' in folder_str.lower():
                match = re.search(r'"([^"]+)"\s*$', folder_str)
                if match:
                    sent_folder = match.group(1)
                    break
                    
    if not sent_folder:
        sent_folder = "[Gmail]/Sent Mail"
        
    mail.select(f'"{sent_folder}"')
    
    yesterday_imap = (datetime.now() - timedelta(days=1)).strftime("%d-%b-%Y")
    status, data = mail.search(None, f'SINCE {yesterday_imap}')
    
    total_sent = 0
    if status == 'OK' and data[0]:
        msg_ids_list = data[0].split()
        if msg_ids_list:
            msg_ids_joined = ",".join(x.decode('utf-8') for x in msg_ids_list)
            fetch_status, fetch_data = mail.fetch(msg_ids_joined, "(INTERNALDATE)")
            
            if fetch_status == 'OK':
                from datetime import timezone
                now_utc = datetime.now(timezone.utc)
                for response_item in fetch_data:
                    item_bytes = None
                    if isinstance(response_item, tuple):
                        item_bytes = response_item[0]
                    elif isinstance(response_item, bytes):
                        item_bytes = response_item
                        
                    if item_bytes:
                        item_str = item_bytes.decode('utf-8', errors='ignore')
                        match = re.search(r'INTERNALDATE "([^"]+)"', item_str)
                        if match:
                            date_str = match.group(1)
                            try:
                                msg_date = datetime.strptime(date_str.strip(), "%d-%b-%Y %H:%M:%S %z")
                                if now_utc - msg_date < timedelta(hours=24):
                                    total_sent += 1
                            except Exception as ex:
                                print(f"DEBUG: Error parsing INTERNALDATE: {ex}")
    mail.logout()
    return total_sent

@router.get("/campaign/emails-sent-today")
async def get_emails_sent_today(spreadsheet_id: str = None, current_user: dict = Depends(get_current_user)):
    from datetime import datetime, timedelta
    
    now_ts = time.time()
    if now_ts - EMAILS_SENT_TODAY_CACHE["timestamp"] < 10.0:
        return {"emails_sent_today": EMAILS_SENT_TODAY_CACHE["count"], "limit": 1000, "source": "cache"}
        
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASS")
    smtp_server = os.getenv("SMTP_SERVER")
    
    if smtp_user and smtp_pass and "gmail" in (smtp_server or "").lower():
        try:
            total_sent = await asyncio.to_thread(fetch_gmail_sent_count, smtp_user, smtp_pass, smtp_server)
            print(f"DEBUG: Successfully fetched direct Gmail rolling 24h count: {total_sent}")
            EMAILS_SENT_TODAY_CACHE["count"] = total_sent
            EMAILS_SENT_TODAY_CACHE["timestamp"] = now_ts
            return {"emails_sent_today": total_sent, "limit": 1000, "source": "gmail_direct"}
        except Exception as e:
            print(f"DEBUG: Gmail IMAP direct fetch failed: {e}. Falling back to sheet logs.")
            
    # 2. Fallback: Count from Google Sheet Logs + Memory (rolling 24 hours)
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID") or spreadsheet_id
    now = datetime.now()
    sheet_counted_contacts = set()
    
    if spreadsheet_id:
        try:
            from app.services.sheet_service import get_sheets_service
            service = get_sheets_service()
            result = await asyncio.to_thread(service.values().get(
                spreadsheetId=spreadsheet_id,
                range="Logs Data!A:H"
            ).execute)
            rows = result.get('values', [])
            
            is_admin = current_user.get("role") == "Admin"
            for row in rows:
                # 12-col layout: [0]=CampaignID [1]=Platform [2]=Name [3]=Company
                # [4]=Phone [5]=Email [6]=Timestamp [7]=Details [8]=GenerateBy
                if len(row) >= 9 and row[0].startswith("CAM") and (is_admin or row[8] == current_user["username"]):
                    if row[1] == "Email" and len(row) > 6:
                        try:
                            row_date = datetime.strptime(str(row[6]).strip(), "%Y-%m-%d %I:%M:%S %p")
                            if now - row_date < timedelta(hours=24):
                                campaign_id = row[0]
                                email = row[5]
                                sheet_counted_contacts.add((campaign_id, email))
                        except Exception:
                            continue
        except Exception as e:
            print(f"DEBUG: Failed to read Logs Data sheet: {e}")
            
    memory_count = 0
    for campaign_id, campaign in campaign_manager.items():
        for res in campaign.get("results", []):
            if res.get("type") == "Email" and res.get("sent_time", ""):
                try:
                    res_date = datetime.strptime(res.get("sent_time"), "%Y-%m-%d %I:%M:%S %p")
                    if now - res_date < timedelta(hours=24):
                        email = res.get("email")
                        key = (campaign_id, email)
                        if key not in sheet_counted_contacts:
                            memory_count += 1
                except Exception:
                    continue
                    
    total_sent = len(sheet_counted_contacts) + memory_count
    EMAILS_SENT_TODAY_CACHE["count"] = total_sent
    EMAILS_SENT_TODAY_CACHE["timestamp"] = now_ts
    return {"emails_sent_today": total_sent, "limit": 1000, "source": "sheet_fallback"}


@router.get("/campaigns")
async def list_campaigns(spreadsheet_id: str = None, current_user: dict = Depends(get_current_user)):
    """Returns a sorted list of unique campaign IDs and their dates created by the current user."""
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID") or spreadsheet_id
    campaign_map = {}
    is_admin = current_user.get("role") == "Admin"
    
    # 1. Filter active in-memory campaigns
    for campaign_id, campaign in campaign_manager.items():
        if is_admin or campaign.get("metadata", {}).get("created_by") == current_user["username"]:
            earliest_date = ""
            for res in campaign.get("results", []):
                if res.get("sent_time"):
                    raw_ts = res.get("sent_time")
                    parsed_date = raw_ts
                    date_formats = [
                        "%Y-%m-%d %I:%M:%S %p",
                        "%Y-%m-%d %H:%M:%S",
                        "%d %B %Y",
                        "%d %B %Y, %I:%M%p",
                        "%d %B %Y, %I:%M %p"
                    ]
                    for date_fmt in date_formats:
                        try:
                            dt = datetime.strptime(raw_ts.strip(), date_fmt)
                            parsed_date = f"{dt.day} {dt.strftime('%B %Y')}"
                            break
                        except ValueError:
                            continue
                    earliest_date = parsed_date
                    break
            campaign_map[campaign_id] = earliest_date
            
    # 2. Filter historical campaigns from sheet
    if spreadsheet_id:
        try:
            from app.services.sheet_service import get_sheets_service
            service = get_sheets_service()
            result = await asyncio.to_thread(service.values().get(
                spreadsheetId=spreadsheet_id,
                range="Logs Data!A:L"
            ).execute)
            values = result.get('values', [])
            for row in values:
                if len(row) >= 2 and row[0].startswith("CAM"):
                    row_generated_by = row[8] if len(row) > 8 else "System"
                    if is_admin or row_generated_by == current_user["username"]:
                        campaign_id = row[0]
                        timestamp_val = row[6] if len(row) > 6 else ""
                        if timestamp_val and campaign_id not in campaign_map:
                            raw_ts = timestamp_val.strip()
                            parsed_date = raw_ts
                            date_formats = [
                                "%Y-%m-%d %I:%M:%S %p",
                                "%Y-%m-%d %H:%M:%S",
                                "%d %B %Y",
                                "%d %B %Y, %I:%M%p",
                                "%d %B %Y, %I:%M %p"
                            ]
                            for date_fmt in date_formats:
                                try:
                                    dt = datetime.strptime(raw_ts, date_fmt)
                                    parsed_date = f"{dt.day} {dt.strftime('%B %Y')}"
                                    break
                                except ValueError:
                                    continue
                            campaign_map[campaign_id] = parsed_date
                        elif campaign_id not in campaign_map:
                            campaign_map[campaign_id] = ""
        except Exception as e:
            print(f"DEBUG: Logs sheet is inaccessible for listing campaigns: {e}")
            
    sorted_ids = sorted(list(campaign_map.keys()), reverse=True)
    return [{"id": cid, "date": campaign_map[cid]} for cid in sorted_ids]


@router.get("/campaign/all/status")
async def get_all_campaigns_status(spreadsheet_id: str = None, current_user: dict = Depends(get_current_user)):
    """Returns log entries from all active and historical campaigns combined, scoped to current user."""
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID") or spreadsheet_id
    all_results = []
    loaded_campaigns = set()
    is_admin = current_user.get("role") == "Admin"
    for campaign_id, campaign in campaign_manager.items():
        if is_admin or campaign.get("metadata", {}).get("created_by") == current_user["username"]:
            for res in campaign.get("results", []):
                if "sent_time" not in res:
                    res["sent_time"] = ""
            all_results.extend(campaign.get("results", []))
            loaded_campaigns.add(campaign_id)
        
    if spreadsheet_id:
        try:
            from app.services.sheet_service import get_sheets_service
            
            # Check cache first
            now = time.time()
            cached = SHEETS_LOGS_CACHE.get(spreadsheet_id)
            if cached and (now - cached["timestamp"] < CACHE_EXPIRY_SECONDS):
                rows = cached["rows"]
            else:
                service = get_sheets_service()
                result = await asyncio.to_thread(service.values().get(
                    spreadsheetId=spreadsheet_id,
                    range="Logs Data!A:L"
                ).execute)
                rows = result.get('values', [])
                SHEETS_LOGS_CACHE[spreadsheet_id] = {
                    "rows": rows,
                    "timestamp": now
                }
            
            for row in rows:
                if len(row) >= 2 and row[0].startswith("CAM"):
                    row_generated_by = row[8] if len(row) > 8 else "System"
                    if is_admin or row_generated_by == current_user["username"]:
                        if row[0] not in loaded_campaigns:
                            platform_val = row[1] if len(row) > 1 else "Email"
                            name_val = row[2] if len(row) > 2 else ""
                            company_val = row[3] if len(row) > 3 else ""
                            phone_val = row[4] if len(row) > 4 else ""
                            email_val = row[5] if len(row) > 5 else ""
                            sent_time_val = row[6] if len(row) > 6 else ""
                            details_val = row[7] if len(row) > 7 else ""
                            
                            status_val = "Sent"
                            if "Seen" in details_val:
                                status_val = "Seen"
                            elif "Failed" in details_val:
                                status_val = "Failed"
                                
                            all_results.append({
                                "campaign_id": row[0],
                                "name": name_val,
                                "company": company_val,
                                "phone": phone_val if platform_val == "WhatsApp" else None,
                                "email": email_val if platform_val == "Email" else None,
                                "type": platform_val,
                                "status": status_val,
                                "reason": details_val,
                                "sent_time": sent_time_val,
                                "generated_by": row_generated_by,
                                "subscription": row[9] if len(row) > 9 else "Yes",
                                "unsub_reason": row[10] if len(row) > 10 else "",
                                "unsub_other": row[11] if len(row) > 11 else ""
                            })
        except Exception as e:
            print(f"DEBUG: Logs sheet is inaccessible for all campaign status polling: {e}")
            
    return all_results


@router.get("/campaign/{campaign_id}/status")
async def get_campaign_status(campaign_id: str, spreadsheet_id: str = None, current_user: dict = Depends(get_current_user)):
    """Returns log entries for a specific campaign ID, scoped to current user, falling back to Google Sheets."""
    is_admin = current_user.get("role") == "Admin"
    if campaign_id in campaign_manager:
        campaign = campaign_manager[campaign_id]
        if is_admin or campaign.get("metadata", {}).get("created_by") == current_user["username"]:
            return campaign.get("results", [])
        return []
        
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID") or spreadsheet_id
    if spreadsheet_id:
        try:
            from app.services.sheet_service import get_sheets_service
            
            # Check cache first
            now = time.time()
            cached = SHEETS_LOGS_CACHE.get(spreadsheet_id)
            if cached and (now - cached["timestamp"] < CACHE_EXPIRY_SECONDS):
                rows = cached["rows"]
            else:
                service = get_sheets_service()
                result = await asyncio.to_thread(service.values().get(
                    spreadsheetId=spreadsheet_id,
                    range="Logs Data!A:L"
                ).execute)
                rows = result.get('values', [])
                SHEETS_LOGS_CACHE[spreadsheet_id] = {
                    "rows": rows,
                    "timestamp": now
                }
            
            results = []
            for row in rows:
                if len(row) >= 2 and row[0] == campaign_id:
                    row_generated_by = row[8] if len(row) > 8 else "System"
                    if is_admin or row_generated_by == current_user["username"]:
                        platform_val = row[1] if len(row) > 1 else "Email"
                        name_val = row[2] if len(row) > 2 else ""
                        company_val = row[3] if len(row) > 3 else ""
                        phone_val = row[4] if len(row) > 4 else ""
                        email_val = row[5] if len(row) > 5 else ""
                        sent_time_val = row[6] if len(row) > 6 else ""
                        details_val = row[7] if len(row) > 7 else ""
                        
                        status_val = "Sent"
                        if "Seen" in details_val:
                            status_val = "Seen"
                        elif "Failed" in details_val:
                            status_val = "Failed"
                            
                        results.append({
                            "campaign_id": row[0],
                            "name": name_val,
                            "company": company_val,
                            "phone": phone_val if platform_val == "WhatsApp" else None,
                            "email": email_val if platform_val == "Email" else None,
                            "type": platform_val,
                            "status": status_val,
                            "reason": details_val,
                            "sent_time": sent_time_val,
                            "generated_by": row_generated_by,
                            "subscription": row[9] if len(row) > 9 else "Yes",
                            "unsub_reason": row[10] if len(row) > 10 else "",
                            "unsub_other": row[11] if len(row) > 11 else ""
                        })
            return results
        except Exception as e:
            print(f"DEBUG: Logs sheet is inaccessible for campaign status lookup: {campaign_id}: {e}")
            
    return []


@router.get("/campaign/{campaign_id}/export")
async def export_campaign_logs(campaign_id: str, spreadsheet_id: str = None, current_user: dict = Depends(get_current_user)):
    """Exports campaign logs for a specific campaign ID to an Excel file."""
    results = await get_campaign_status(campaign_id, spreadsheet_id, current_user)
    
    if not results:
        raise HTTPException(status_code=404, detail=f"No logs found for campaign {campaign_id}")
        
    # Parse starting SNO from campaign ID if present (e.g. CAM002(89-91))
    import re
    start_sno = None
    range_match = re.search(r"\((\d+)-(\d+)\)", campaign_id)
    if range_match:
        try:
            start_sno = int(range_match.group(1))
        except ValueError:
            pass

    export_data = []
    for idx, item in enumerate(results, start=1):
        # Use original SNO from results if available, otherwise calculate from campaign ID range
        sno_val = item.get("sno")
        if not sno_val:
            if start_sno is not None:
                sno_val = str(start_sno + idx - 1)
            else:
                sno_val = str(idx)
        
        # Parse and format the Date column to "DD/MM/YYYY Time"
        raw_date = item.get("sent_time", "")
        formatted_date = raw_date
        if raw_date:
            # Common formats to try
            date_formats = [
                "%Y-%m-%d %I:%M:%S %p",
                "%Y-%m-%d %H:%M:%S",
                "%d %B %Y, %I:%M%p",
                "%d %B %Y, %H:%M",
                "%d %B %Y, %I:%M %p"
            ]
            for date_fmt in date_formats:
                try:
                    dt = datetime.strptime(raw_date.strip(), date_fmt)
                    formatted_date = dt.strftime("%d/%m/%Y %I:%M:%S %p")
                    break
                except ValueError:
                    continue

        export_data.append({
            "SNO": sno_val,
            "Campaign ID": campaign_id,
            "Date": formatted_date,
            "Name": item.get("name", ""),
            "Email": item.get("email") or "",
            "Phone": item.get("phone") or "",
            "Status": item.get("status", ""),
            "Detail": item.get("reason", ""),
            "Sent By": item.get("generated_by", ""),
            "Subscription (Yes / No)": item.get("subscription", "Yes"),
            "Unsubscribe Reason": item.get("unsub_reason", ""),
            "If Other (Reason)": item.get("unsub_other", "")
        })
        
    df = pd.DataFrame(export_data)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=campaign_id, index=False)
        
        # Access openpyxl workbook and worksheet to apply styling
        workbook = writer.book
        worksheet = writer.sheets[campaign_id]
        
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # Apply style to all header cells (row 1)
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            
        # Auto-adjust column widths based on maximum content length
        for col_idx in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(df.columns[col_idx - 1])
            for val in df.iloc[:, col_idx - 1]:
                max_len = max(max_len, len(str(val or "")))
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    buffer.seek(0)
    
    filename = f"{campaign_id}_logs.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )


@router.get("/campaign/track-open")
async def track_open(campaign_id: str, email: str):
    from datetime import datetime
    print(f"DEBUG: Track open received for campaign {campaign_id}, email {email}")
    
    opened_time = datetime.now().strftime("%I:%M %p")
    details = f"Seen at {opened_time}"
    
    # 1. Update in-memory if available
    spreadsheet_id = None
    if campaign_id in campaign_manager:
        campaign = campaign_manager[campaign_id]
        spreadsheet_id = campaign.get("metadata", {}).get("spreadsheet_id")
        for item in campaign.get("results", []):
            if item.get("email") == email and item.get("status") != "Seen":
                item["status"] = "Seen"
                item["reason"] = details
                break
                
    # 2. Always fall back or default to LOGS_SPREADSHEET_ID from env
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID") or spreadsheet_id
    
    # 3. Always update Google Sheets directly if spreadsheet_id is resolved
    if spreadsheet_id:
        try:
            print(f"DEBUG: Enqueueing Google Sheets log update to 'Seen' for campaign {campaign_id}, email {email}")
            asyncio.create_task(update_log_status(spreadsheet_id, campaign_id, email, "Seen", details))
        except Exception as e:
            print(f"DEBUG: Failed to update Logs Data sheet on open tracking: {e}")

    gif_bytes = b'GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
    return Response(
        content=gif_bytes,
        media_type="image/gif",
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0"
        }
    )



class UnsubscribeRequest(BaseModel):
    email: str
    campaign_id: str = ""
    reason: str = ""
    other_reason: str = ""

@router.post("/campaign/unsubscribe")
async def handle_unsubscribe(req: UnsubscribeRequest):
    """Handles unsubscribe requests from the unsubscribe page. No auth required."""
    print(f"DEBUG: Unsubscribe request received for email={req.email}, campaign_id={req.campaign_id}, reason={req.reason}")
    try:
        updated = await update_unsubscribe_status(
            spreadsheet_id=None,  # reads from LOGS_SPREADSHEET_ID env var
            email=req.email,
            campaign_id=req.campaign_id,
            reason=req.reason,
            other_reason=req.other_reason
        )
        # Invalidate the log cache so the frontend fetches fresh data immediately
        SHEETS_LOGS_CACHE.clear()
        if updated:
            return {"success": True, "message": "Unsubscribed successfully."}
        else:
            # Email not found in logs — still succeed for UX, just log it
            print(f"DEBUG: Email {req.email} not found in Logs Data sheet for unsubscribe update.")
            return {"success": True, "message": "Unsubscribed. (No matching log entry found to update.)"}
    except Exception as e:
        print(f"DEBUG: Unsubscribe handler error: {e}")
        raise HTTPException(status_code=500, detail="Failed to process unsubscribe request.")

@router.post("/rewrite")
async def ai_rewrite(message: str = Form(...), tone: str = Form("professional"), current_user: dict = Depends(get_current_user)):
    rewritten = await rewrite_message(message, tone)
    return {"rewritten": rewritten}

# Schemas for User CRUD
class UserCreateRequest(BaseModel):
    id: str
    username: str
    password: str
    role: str

class UserUpdateRequest(BaseModel):
    username: str
    password: str
    role: str

@router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Admin":
        raise HTTPException(status_code=403, detail="Forbidden: Admin access required.")
        
    users_rows = await read_users_data(None)
    if not users_rows or len(users_rows) <= 1:
        return []
        
    headers = [str(h).strip().lower() for h in users_rows[0]]
    try:
        id_idx = headers.index("id")
        username_idx = headers.index("user name")
        password_idx = headers.index("password")
        role_idx = headers.index("role") if "role" in headers else -1
    except ValueError:
        raise HTTPException(
            status_code=500,
            detail="Users spreadsheet is missing required columns."
        )
        
    users = []
    for row in users_rows[1:]:
        if len(row) > max(id_idx, username_idx, password_idx):
            u_id = str(row[id_idx]).strip()
            u_name = str(row[username_idx]).strip()
            u_pass = str(row[password_idx]).strip()
            u_role = "User"
            if role_idx != -1 and len(row) > role_idx:
                u_role = str(row[role_idx]).strip()
            users.append({
                "id": u_id,
                "username": u_name,
                "password": u_pass,
                "role": u_role
            })
    return users

@router.post("/users")
async def create_user(req: UserCreateRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Admin":
        raise HTTPException(status_code=403, detail="Forbidden: Admin access required.")
        
    users_rows = await read_users_data(None)
    
    # Check for duplicate user ID or username
    for row in users_rows[1:]:
        if len(row) > 1:
            if str(row[0]).strip().lower() == req.id.strip().lower():
                raise HTTPException(status_code=400, detail="User ID already exists.")
            if str(row[1]).strip().lower() == req.username.strip().lower():
                raise HTTPException(status_code=400, detail="Username already exists.")
                
    new_row = [req.id.strip(), req.username.strip(), req.password.strip(), req.role.strip()]
    
    from app.services.sheet_service import get_sheets_service
    import os
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID")
    if not spreadsheet_id:
        raise HTTPException(status_code=500, detail="Logs spreadsheet ID not configured.")
        
    service = get_sheets_service()
    await asyncio.to_thread(service.values().append(
        spreadsheetId=spreadsheet_id,
        range="Users!A:D",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": [new_row]}
    ).execute)
    
    return {"message": "User created successfully"}

@router.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdateRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Admin":
        raise HTTPException(status_code=403, detail="Forbidden: Admin access required.")
        
    users_rows = await read_users_data(None)
    if not users_rows:
        raise HTTPException(status_code=404, detail="Users sheet is empty.")
        
    updated = False
    new_rows = [users_rows[0]]  # headers
    
    for row in users_rows[1:]:
        if len(row) > 0 and str(row[0]).strip().lower() == user_id.strip().lower():
            # Update values
            # Ensure length is 4
            updated_row = [str(row[0]).strip(), req.username.strip(), req.password.strip(), req.role.strip()]
            new_rows.append(updated_row)
            updated = True
        else:
            new_rows.append(row)
            
    if not updated:
        raise HTTPException(status_code=404, detail="User not found.")
        
    from app.services.sheet_service import get_sheets_service
    import os
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID")
    if not spreadsheet_id:
        raise HTTPException(status_code=500, detail="Logs spreadsheet ID not configured.")
        
    service = get_sheets_service()
    # Clear and rewrite
    await asyncio.to_thread(service.values().clear(
        spreadsheetId=spreadsheet_id,
        range="Users!A1:D200"
    ).execute)
    
    await asyncio.to_thread(service.values().update(
        spreadsheetId=spreadsheet_id,
        range=f"Users!A1:D{len(new_rows)}",
        valueInputOption="RAW",
        body={'values': new_rows}
    ).execute)
    
    return {"message": "User updated successfully"}

@router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Admin":
        raise HTTPException(status_code=403, detail="Forbidden: Admin access required.")
        
    users_rows = await read_users_data(None)
    if not users_rows:
        raise HTTPException(status_code=404, detail="Users sheet is empty.")
        
    deleted = False
    new_rows = [users_rows[0]]  # headers
    
    for row in users_rows[1:]:
        if len(row) > 0 and str(row[0]).strip().lower() == user_id.strip().lower():
            deleted = True
        else:
            new_rows.append(row)
            
    if not deleted:
        raise HTTPException(status_code=404, detail="User not found.")
        
    from app.services.sheet_service import get_sheets_service
    import os
    spreadsheet_id = os.getenv("LOGS_SPREADSHEET_ID")
    if not spreadsheet_id:
        raise HTTPException(status_code=500, detail="Logs spreadsheet ID not configured.")
        
    service = get_sheets_service()
    # Clear and rewrite
    await asyncio.to_thread(service.values().clear(
        spreadsheetId=spreadsheet_id,
        range="Users!A1:D200"
    ).execute)
    
    await asyncio.to_thread(service.values().update(
        spreadsheetId=spreadsheet_id,
        range=f"Users!A1:D{len(new_rows)}",
        valueInputOption="RAW",
        body={'values': new_rows}
    ).execute)
    
    return {"message": "User deleted successfully"}
